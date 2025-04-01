from openpyxl import load_workbook
# Load the workbook
wb = load_workbook('your_file.xlsx')
# Get a sheet
sheet = wb['Sheet1']  # or wb.active for the active sheet
# Read cell values
for row in sheet.iter_rows(values_only=True):
    print(row)
